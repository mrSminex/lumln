"""
Экспорт клиентской базы в Excel (.xlsx).

Создаёт книгу с двумя листами:
  - «Клиенты»  — все клиенты с количеством формул
  - «Формулы»  — все формулы с привязкой к клиенту
"""

import io
from datetime import datetime

import aiosqlite
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from config import DB_PATH


# ─── Стили ────────────────────────────────────────────────────────────────────

DARK   = "2D2D2D"   # почти чёрный — фирменный цвет LUM'N
BEIGE  = "F5E6D3"   # бежевый — второй фирменный цвет
WHITE  = "FFFFFF"
LIGHT  = "FAF9F7"   # очень светлый, чётные строки

_header_font  = Font(bold=True, color=WHITE,  size=10)
_header_fill  = PatternFill("solid", fgColor=DARK)
_alt_fill     = PatternFill("solid", fgColor=LIGHT)
_title_font   = Font(bold=True, color=DARK, size=13)
_subtitle     = Font(color="888888", size=9)

_thin = Side(style="thin", color="E0DDD8")
_border = Border(bottom=_thin)

_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _header_row(ws, row: int, columns: list[tuple[str, int]]) -> None:
    """Рисует строку заголовков с тёмным фоном."""
    for col_idx, (title, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _center
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _data_row(ws, row: int, values: list, alt: bool = False) -> None:
    fill = _alt_fill if alt else None
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.alignment = _left
        cell.border = _border
        if fill:
            cell.fill = fill


# ─── Лист «Клиенты» ───────────────────────────────────────────────────────────

async def _sheet_clients(wb: openpyxl.Workbook) -> int:
    ws = wb.active
    ws.title = "Клиенты"
    ws.sheet_view.showGridLines = False

    # Заголовок документа
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"LUM'N — База клиентов   (выгружено {datetime.now().strftime('%d.%m.%Y %H:%M')})"
    title_cell.font = _title_font
    title_cell.alignment = _center
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")  # пустая строка-отступ

    columns = [
        ("№ клиента", 12),
        ("Имя",       28),
        ("Телефон",   18),
        ("Формул",    10),
        ("Telegram ID", 18),
        ("Дата регистрации", 20),
    ]
    _header_row(ws, row=3, columns=columns)
    ws.row_dimensions[3].height = 22

    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("""
            SELECT c.id, c.name, c.phone, c.telegram_id, c.created_at,
                   COUNT(f.id) AS formula_count
            FROM clients c
            LEFT JOIN formulas f ON f.client_id = c.id
            GROUP BY c.id
            ORDER BY c.id
        """) as cur:
            clients = [dict(r) for r in await cur.fetchall()]

    for i, c in enumerate(clients, start=1):
        _data_row(ws, row=i + 3, values=[
            c["id"],
            c["name"],
            c["phone"],
            c["formula_count"],
            c["telegram_id"],
            c["created_at"][:16],
        ], alt=(i % 2 == 0))

    # Freeze header
    ws.freeze_panes = "A4"
    return len(clients)


# ─── Лист «Формулы» ───────────────────────────────────────────────────────────

async def _sheet_formulas(wb: openpyxl.Workbook) -> int:
    ws = wb.create_sheet("Формулы")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"LUM'N — Все формулы   (выгружено {datetime.now().strftime('%d.%m.%Y %H:%M')})"
    title_cell.font = _title_font
    title_cell.alignment = _center
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")

    columns = [
        ("№ формулы", 12),
        ("№ клиента", 12),
        ("Имя клиента", 26),
        ("Телефон",    18),
        ("Название формулы", 32),
        ("Состав",          48),
        ("Дата создания",   20),
    ]
    _header_row(ws, row=3, columns=columns)
    ws.row_dimensions[3].height = 22

    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("""
            SELECT f.id, f.client_id, f.title, f.content, f.created_at,
                   c.name AS client_name, c.phone
            FROM formulas f
            JOIN clients c ON c.id = f.client_id
            ORDER BY f.client_id, f.created_at
        """) as cur:
            formulas = [dict(r) for r in await cur.fetchall()]

    for i, f in enumerate(formulas, start=1):
        row_num = i + 3
        _data_row(ws, row=row_num, values=[
            f["id"],
            f["client_id"],
            f["client_name"],
            f["phone"],
            f["title"],
            f["content"],
            f["created_at"][:16],
        ], alt=(i % 2 == 0))
        # Высокие строки для состава формулы
        ws.row_dimensions[row_num].height = max(
            15, min(80, f["content"].count("\n") * 14 + 14)
        )

    ws.freeze_panes = "A4"
    return len(formulas)


# ─── Публичные функции ───────────────────────────────────────────────────────

async def build_excel_full() -> tuple[bytes, str]:
    """Два листа: Клиенты + Формулы."""
    wb = openpyxl.Workbook()
    await _sheet_clients(wb)
    await _sheet_formulas(wb)
    return _save(wb, "lumln_full")


async def build_excel_clients() -> tuple[bytes, str]:
    """Один лист: только Клиенты."""
    wb = openpyxl.Workbook()
    await _sheet_clients(wb)
    return _save(wb, "lumln_clients")


async def build_excel_formulas() -> tuple[bytes, str]:
    """Один лист: только Формулы."""
    wb = openpyxl.Workbook()
    # openpyxl создаёт пустой Sheet по умолчанию — переименуем его под формулы
    wb.active.title = "_tmp"
    await _sheet_formulas(wb)
    del wb["_tmp"]
    return _save(wb, "lumln_formulas")


def _save(wb: openpyxl.Workbook, prefix: str) -> tuple[bytes, str]:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return buf.read(), filename
